# Selenium imports
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

# Openpyxl imports
from openpyxl import load_workbook

# Operating system imports
import os

# Todos
# TODO: Make program accept xlsx documents and Parse for phone number and photo path


def xl_parser():
    os.chdir('/Users/alexharris/Desktop/')
    filename = raw_input("What file would you like to Parse? [Copy absolute path to file here]: ")

    wb = load_workbook(filename)

    sheetlist = []
    sheetlist.append(wb.sheetnames)
    sheet = wb['Sheet1']
    total_rows = sheet.max_row

    # Print sheets in file for user
    for sheet in range(len(sheetlist)):
        print("Sheets available in file: " + str(sheetlist[sheet]) + '\n')

    # Store sheetlist as set and check against user choice
    # TODO: Make this work
    # sheet_choice = raw_input("Which sheet would you like to load into the parser? [type indexno. ] ")
    # if sheet_choice in sheetlist:
    #     sheet = wb[sheet_choice]
    # else:
    #     print("Sheet not found")

    # Get all available columns and display vertically
    # TODO: Fix this to properly display values
    available_columns = []
    sheet = wb['Sheet1']
    for cellObj in sheet['A1':'B1']:
        for cells in cellObj:
            available_columns.extend([cells.coordinate, cells.value])
            for i in available_columns:
                print (i + '\n')

    counter = 0
    number_list = []
    photo_list = []
    while counter == 0:
        columnSelect1 = raw_input("Which column contains your numbers? ").upper()  # Make uppercase value for proper parameter use later
        columnSelect2 = raw_input("Which column contains your photos ").upper()

        sheet = wb['Sheet1']
        total_rows = sheet.max_row
        for cellObj in sheet[columnSelect1:columnSelect1[0] + str(total_rows)]:
            for cells in cellObj:
                number_list.append(cells.value)

        for cellObj in sheet[columnSelect2:columnSelect2[0] + str(total_rows)]:
            for cells in cellObj:
                photo_list.append(cells.value)
        counter = 1
    print(str(total_rows) + " / " + str(len(number_list)) + " items populated to numbers list")
    print(str(total_rows) + " / " + str(len(photo_list)) + " items populated to photos list")
    merged_list = zip(number_list, photo_list) # Creates tuple list in format: (number, photo)
    return merged_list


def main():
    number_list = []
    photo_list = []
    for item in xl_parser():
        unpacked_attendee = item
        (number, photo) = unpacked_attendee
        number_list.append(number)
        photo_list.append(photo)

    driver = webdriver.Chrome(executable_path="/Users/alexharris/Desktop/chromedriver")
    driver.get('https://web.whatsapp.com/')

    start_automation = raw_input("Please scan QR code in browser [Press any key to continue] ")

    if start_automation:
        # Edit these fields to dynamically change the contact and msg
        #  TODO: Put this code in a loop for r in total_rows
        counter = 0
        while counter != len(number_list):
            for number in number_list:
                contact_id = number
            for photo in photo_list:
                msg_content = str(photo)

            # Focus search_bar and bring up contact
                search_bar_container_id = driver.find_element_by_css_selector('#side > div._3CPl4 > div > label > input')
                search_bar_container_id.click()
                search_bar_container_id.send_keys(contact_id)
                search_bar_container_id.send_keys(Keys.ENTER)

            # Focus message box and send msg_content
                msg_box = driver.find_element_by_css_selector('#main > footer > div._3pkkz.copyable-area > div._1Plpp > div > div._2S1VP.copyable-text.selectable-text')
                msg_box.click()
                msg_box.send_keys(msg_content)
                msg_box.send_keys(Keys.ENTER)
        counter = counter + 1

if __name__ == "__main__":
    main()
